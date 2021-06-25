from flask import Flask, flash, render_template, url_for, request, redirect, session
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
import random
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, login_manager, current_user
from werkzeug.security import check_password_hash, generate_password_hash
import openpyxl


from flask import Flask
from flask_login import LoginManager
from flask_migrate import Migrate
from flask_sqlalchemy import SQLAlchemy

app = Flask(__name__)
manager = LoginManager(app)

app.secret_key = 'some secret key'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///app.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

migrate = Migrate(app, db)
db.create_all()
class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    login = db.Column(db.String(100), nullable=False, unique=True)
    password = db.Column(db.String(255), nullable=False)
    words = db.relationship('Cidian', backref = 'author', lazy = 'dynamic')

    def __repr__(self):
        return '<User %r>' % self.id


class Cidian(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    eng = db.Column(db.String(100), nullable=False)
    sp = db.Column(db.String(100), nullable=False)
    img_url = db.Column(db.String(100), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'))

    def __repr__(self):
        return self.id

class Globls:
    all_my_words = []
    total = 0
    score = 0
    wrong_ans = {}
    eng = []
    names = []
    maxs = 0
    img_url = []
    sp = []
    viewed = 0



@app.route('/', methods=['GET','POST'])
@login_required
def index():
    df0 = openpyxl.load_workbook('all_words.xlsx')
    mas = df0.sheetnames
    if request.method == "POST":
        a = request.form.get('refer')
        return redirect('/to_'+str(a))
    return render_template('index.html', mas=mas)


@app.route('/my_wall', methods=['GET','POST'])
@login_required
def my_wall():
    cur_user = User.query.filter_by(id=current_user.id).first()
    my_words = cur_user.words
    my_eng = 0
    if request.method == "POST":
        my_word = request.form.get('word')
        d = request.form.get('delete')
        if d:
            word = Cidian.query.filter_by(eng=d).first()
            db.session.delete(word)
            db.session.commit()
            Globls.all_my_words = []
        if my_word:
            cur_user = User.query.filter_by(id=current_user.id).first()
            my_words = cur_user.words
            my_eng = Cidian.query.filter_by(eng=my_word).first()
        h = request.form.get('hear')
        if h:
            cur_user = User.query.filter_by(id=current_user.id).first()
            my_words = cur_user.words
            my_eng = Cidian.query.filter_by(eng=h).first()
        return render_template('my_wall.html', eng=my_eng, h=h, my_words=my_words)

    return render_template('my_wall.html', my_words=my_words)


@app.route('/my_wall/<string:my_word>', methods=['GET','POST'])
def in_my_wall(my_word):
    cur_user = User.query.filter_by(id=current_user.id).first()
    my_words = cur_user.words
    my_eng = Cidian.query.filter_by(eng=my_word).first()

    return render_template('my_wall.html', eng=my_eng, my_words=my_words)


@app.route('/review', methods=['GET', 'POST'])
def review():
    if (len(Globls.all_my_words)==0 and Globls.total==0):
        cur_user = User.query.filter_by(id=current_user.id).first()
        my_words = cur_user.words
        for el in my_words:
            Globls.all_my_words.append(el)
        random.shuffle(Globls.all_my_words)

    if(len(Globls.all_my_words)==0 and Globls.total==0):
        return render_template('review.html',total = 0)

    if request.method == "POST":
        Globls.all_my_words.pop()
        translate = request.form.get('input')
        Globls.total += 1
        word = request.form.get('word')
        if word.lower() == translate:
            Globls.score += 1
        else:
            img = Cidian.query.filter_by(sp=word).first()
            img_url = img.img_url
            Globls.wrong_ans[word] = img_url

        if (len(Globls.all_my_words) == 0 and Globls.total != 0):
            sc = Globls.score
            tot = Globls.total
            w_a = Globls.wrong_ans
            Globls.score = 0
            Globls.total = 0
            Globls.wrong_ans = {}
            Globls.all_my_words = []
            return render_template('review.html', score=sc, total=tot, all_my_words=Globls.all_my_words, wrong_ans=w_a)

    if (len(Globls.all_my_words) > 0):
        el = Globls.all_my_words[len(Globls.all_my_words) - 1]
        return render_template('review.html', el=el, score=Globls.score, total=Globls.total,all_my_words=Globls.all_my_words)


@app.route('/<string:name>', methods=['GET', 'POST'])
def to_vocab(name):
    if name[:2]=="to" and (len(Globls.names) == 0 or Globls.names[len(Globls.names)-1] != name):
        Globls.viewed = 0
        Globls.sp = []
        Globls.eng = []
        Globls.img_url = []
        print(name[:2])
        if len(Globls.names) > 0:
            Globls.names.pop()
        Globls.names.append(name)

    if (Globls.viewed==0):
        Globls.sp = []
        Globls.img_url = []
        Globls.eng = []
        df0 = openpyxl.load_workbook('all_words.xlsx')
        cur_df = df0[name[3:]]
        count_row = cur_df.max_row
        for i in range(1, count_row + 1):
            Globls.eng.append(cur_df.cell(row=i, column=1).value)
            Globls.img_url.append(cur_df.cell(row=i, column=3).value)
            Globls.sp.append(cur_df.cell(row=i, column=2).value)
        Globls.maxs = len(Globls.eng)

    word_eng = Globls.eng[Globls.viewed]
    word_sp = Globls.sp[Globls.viewed]
    word_img = Globls.img_url[Globls.viewed]

    if request.method == "POST":
        b = request.form.get('add')
        p = request.form.get('prev')
        n = request.form.get('next')
        h = request.form.get('hear')
        if p:
            Globls.viewed -= 1
            word_eng = Globls.eng[Globls.viewed]
            word_sp = Globls.sp[Globls.viewed]
            word_img = Globls.img_url[Globls.viewed]
            return render_template('to_vocab.html', eng=word_eng, sp=word_sp, img=word_img, name=name, viewed=Globls.viewed,
                                   maxs=Globls.maxs)

        if n:
            Globls.viewed += 1
            print(Globls.viewed)
            word_eng = Globls.eng[Globls.viewed]
            word_sp = Globls.sp[Globls.viewed]
            word_img = Globls.img_url[Globls.viewed]
            print("rrrrrr")
            print(word_eng)
            return render_template('to_vocab.html', eng=word_eng, sp=word_sp, img=word_img, name=name, viewed=Globls.viewed,maxs=Globls.maxs)

        if b:
            if current_user:
                find = Cidian.query.filter_by(eng=b).first()
                if not find:
                    new_word = Cidian(eng=word_eng, sp=word_sp, img_url=word_img, author=current_user)
                    db.session.add(new_word)
                    db.session.commit()
                return render_template('to_vocab.html', eng=word_eng, sp=word_sp, img=word_img, name=name, viewed = Globls.viewed, maxs = Globls.maxs)
        if h:
            return render_template('to_vocab.html',h = h, eng=word_eng, sp=word_sp, img=word_img, name=name, viewed = Globls.viewed, maxs = Globls.maxs)
    return render_template('to_vocab.html',eng = word_eng, sp = word_sp, img = word_img, name = name, viewed = Globls.viewed, maxs = Globls.maxs)


@app.route('/login', methods=['GET', 'POST'])
def login_page():
    login = request.form.get('login')
    password = request.form.get('password')

    if login and password:
        user = User.query.filter_by(login=login).first()

        if user and check_password_hash(user.password, password):
            login_user(user)
            return redirect('/')
        else:
            flash('Login or password is not correct')
    else:
        flash('Please fill login and password fields')

    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    login = request.form.get('login')
    password = request.form.get('password')
    password2 = request.form.get('password2')

    if request.method == 'POST':
        if not (login or password or password2):
            flash('Please, fill all fields!')
        elif password != password2:
            flash('Passwords are not equal!')
        elif db.session.query(User.login).filter_by(login=login).scalar() is  None:

            hash_pwd = generate_password_hash(password)
            new_user = User(login=login, password=hash_pwd)
            db.session.add(new_user)
            db.session.commit()

            return redirect(url_for('login_page'))
        else:
            flash('login already exists')
    else:
        flash('Please, fill login and passport fields')
    return render_template('register.html')


@app.route('/logout', methods=['GET', 'POST'])
@login_required
def logout():
    logout_user()
    return redirect('/login')


@app.after_request
def redirect_to_signin(response):
    if response.status_code == 401:
        return redirect(url_for('login_page') + '?next=' + request.url)
    return response


@manager.user_loader
def load_user(user_id):
    return User.query.get(user_id)


if __name__ == "__main__":
    app.run(debug=True)

